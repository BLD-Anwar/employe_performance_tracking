import os
import csv
import json
import datetime
from database import db_cursor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_report(snapshot_data, meeting_rows, history_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Task Performance Report"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_family = "Segoe UI"
    
    # Colors
    color_dark_green = "004D40"   # Header banner background
    color_teal = "00796B"         # Section title background
    color_zebra_1 = "FFFFFF"       # Zebra row 1
    color_zebra_2 = "F0F7F6"       # Zebra row 2
    color_border = "D3D3D3"        # Light gray borders
    
    # Fonts
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_label_bold = Font(name=font_family, size=10, bold=True)
    font_data = Font(name=font_family, size=10)
    font_italic = Font(name=font_family, size=10, italic=True)
    
    # Fills
    fill_title = PatternFill(start_color=color_dark_green, end_color=color_dark_green, fill_type="solid")
    fill_section = PatternFill(start_color=color_teal, end_color=color_teal, fill_type="solid")
    fill_header = PatternFill(start_color=color_dark_green, end_color=color_dark_green, fill_type="solid")
    fill_zebra_1 = PatternFill(start_color=color_zebra_1, end_color=color_zebra_1, fill_type="solid")
    fill_zebra_2 = PatternFill(start_color=color_zebra_2, end_color=color_zebra_2, fill_type="solid")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    align_title = Alignment(horizontal="center", vertical="center")
    
    # Borders
    border_thin = Side(border_style="thin", color=color_border)
    border_data = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # Row Heights
    ws.row_dimensions[1].height = 40
    
    # A. Title Banner (Merged A1 to D1)
    ws.merge_cells("A1:D1")
    cell_a1 = ws["A1"]
    cell_a1.value = "SRI SRI SUGAR — TASK PERFORMANCE REPORT"
    cell_a1.font = font_title
    cell_a1.fill = fill_title
    cell_a1.alignment = align_title
    
    # Apply styling to merged title cells
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = fill_title
        
    ws.append([]) # Row 2 (empty spacing)
    ws.row_dimensions[2].height = 15
    
    # B. Metadata / Task Details
    task = snapshot_data.get("task", {})
    metrics = snapshot_data.get("metrics", {})
    kpis = snapshot_data.get("kpis", {})
    ranking = snapshot_data.get("ranking", {})
    
    metadata = [
        ("Task ID", task.get("task_id", "")),
        ("Task Name", task.get("task_name", "")),
        ("Work Type", task.get("work_type", "")),
        ("Schedule", task.get("schedule_type", "")),
        ("Priority", task.get("priority", "")),
        ("Assigned Officer", f"{task.get('officer_name', '')} ({task.get('officer_code', '')})"),
        ("Start Date", task.get("start_date", "")),
        ("End Date", task.get("end_date", "")),
        ("Task Status", task.get("status", "")),
        ("Completion Rate", f"{metrics.get('completion_rate', 0.0)}% ({metrics.get('completed_farmers', 0)}/{metrics.get('farmer_count', 0)} Farmers)"),
        ("Meetings Coverage", f"{kpis.get('meetings_coverage_pct', 0.0)}%"),
        ("Calculated Score", f"{kpis.get('calculated_score', 0.0)}% (Grade: {kpis.get('grade', 'N/A')})"),
        ("Officer Rank", f"#{ranking.get('rank', 0)} of {ranking.get('total_officers', 0)}"),
    ]
    
    current_row = 3
    for label, val in metadata:
        # Col A: Label
        ws.cell(row=current_row, column=1, value=label).font = font_label_bold
        ws.cell(row=current_row, column=1).alignment = align_right
        ws.cell(row=current_row, column=1).border = border_data
        
        # Col B: Value (Merged B, C, D)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        cell_val = ws.cell(row=current_row, column=2, value=val)
        cell_val.font = font_data
        cell_val.alignment = align_left
        cell_val.border = border_data
        
        for col in range(2, 5):
            ws.cell(row=current_row, column=col).border = border_data
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    ws.append([]) # Empty spacer row
    current_row += 1
    
    # C. Farmer Meeting Details
    num_cols_meeting = 22
    col_letter_meeting_end = get_column_letter(num_cols_meeting)
    
    ws.merge_cells(f"A{current_row}:{col_letter_meeting_end}{current_row}")
    cell_sec1 = ws.cell(row=current_row, column=1, value="=== FARMER MEETING DETAILS ===")
    cell_sec1.font = font_section
    cell_sec1.fill = fill_section
    cell_sec1.alignment = align_title
    
    for col in range(1, num_cols_meeting + 1):
        ws.cell(row=current_row, column=col).fill = fill_section
        
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    meeting_headers = [
        "Farmer Code", "Farmer Name", "Village",
        "KCC", "KCC Reason",
        "Canara HNT", "Canara Reason",
        "Sangola HNT", "Sangola Reason",
        "Cane Registration", "Cane Reg Remark",
        "Recovery", "Recovery Reason",
        "Vehicle Agreement", "Vehicle Reason",
        "Expected Tonnage", "Cane Development",
        "Feedback", "Remark", "Meeting Date",
        "Vehicle Working", "Vehicle Working Reason"
    ]
    
    for col_idx, header_text in enumerate(meeting_headers, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=header_text)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_data
        
    ws.row_dimensions[current_row].height = 28
    current_row += 1
    
    def yn(val):
        if val is None: return "—"
        return "Yes" if val else "No"
        
    for r_idx, r in enumerate(meeting_rows):
        row_data = [
            r[0], r[1], r[2],
            yn(r[3]), r[4] or "",
            yn(r[5]), r[6] or "",
            yn(r[7]), r[8] or "",
            yn(r[9]), r[10] or "",
            yn(r[11]), r[12] or "",
            yn(r[13]), r[14] or "",
            r[15] or 0, r[16] or "",
            r[17] or "", r[18] or "", r[19] or "",
            yn(r[20]), r[21] or ""
        ]
        
        fill_row = fill_zebra_2 if r_idx % 2 == 1 else fill_zebra_1
        
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = font_data
            c.fill = fill_row
            c.border = border_data
            
            if col_idx in [4, 6, 8, 10, 12, 14, 20]:
                c.alignment = align_center
            elif col_idx in [1, 16, 21]:
                c.alignment = align_center
            else:
                c.alignment = align_left
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    ws.append([]) # Spacer row
    current_row += 1
    
    # D. Task Activity Timeline / History
    num_cols_history = 4
    col_letter_history_end = get_column_letter(num_cols_history)
    
    ws.merge_cells(f"A{current_row}:{col_letter_history_end}{current_row}")
    cell_sec2 = ws.cell(row=current_row, column=1, value="=== TASK ACTIVITY TIMELINE / HISTORY ===")
    cell_sec2.font = font_section
    cell_sec2.fill = fill_section
    cell_sec2.alignment = align_title
    
    for col in range(1, num_cols_history + 1):
        ws.cell(row=current_row, column=col).fill = fill_section
        
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    history_headers = ["Timestamp", "Action", "Remarks", "Officer / Manager"]
    for col_idx, header_text in enumerate(history_headers, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=header_text)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_data
        
    ws.row_dimensions[current_row].height = 28
    current_row += 1
    
    if not history_rows:
        ws.merge_cells(f"A{current_row}:{col_letter_history_end}{current_row}")
        c = ws.cell(row=current_row, column=1, value="No lifecycle logs recorded for this task.")
        c.font = font_italic
        c.alignment = align_center
        c.border = border_data
        
        for col in range(1, num_cols_history + 1):
            ws.cell(row=current_row, column=col).border = border_data
            
        ws.row_dimensions[current_row].height = 20
        current_row += 1
    else:
        for r_idx, hr in enumerate(history_rows):
            row_data = [hr[0], hr[1], hr[2], (hr[3] or "").strip() or "System"]
            fill_row = fill_zebra_2 if r_idx % 2 == 1 else fill_zebra_1
            
            for col_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font = font_data
                c.fill = fill_row
                c.border = border_data
                
                if col_idx == 1:
                    c.alignment = align_center
                else:
                    c.alignment = align_left
                    
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
    # E. Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row in [1] or (cell.value and str(cell.value).startswith("===")):
                continue
            if cell.value:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    return wb


# Base directory for storing static report files
REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")


def get_dynamic_status(status: str, end_date) -> str:
    """Helper to calculate status on the fly based on current date."""
    if not end_date:
        return status

    if isinstance(end_date, str):
        try:
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            try:
                end_date = datetime.datetime.strptime(end_date.split()[0], "%Y-%m-%d").date()
            except ValueError:
                return status
    elif isinstance(end_date, datetime.datetime):
        end_date = end_date.date()
    elif isinstance(end_date, datetime.date):
        pass
    else:
        return status

    today = datetime.date.today()
    if end_date < today and status not in ('COMPLETED', 'CANCELLED', 'OVERDUE'):
        return 'OVERDUE'
    return status


def compute_grade(score):
    """Return letter grade based on score."""
    if score >= 95:
        return "A+"
    elif score >= 85:
        return "A"
    elif score >= 75:
        return "B"
    elif score >= 60:
        return "C"
    elif score >= 45:
        return "D"
    else:
        return "F"


def compute_officer_rankings():
    """
    Compute a ranking of all officers based on their meeting KPI scores.
    Returns dict: {officer_id: {"rank": int, "total": int, "score": float}}
    """
    with db_cursor() as cur:
        # Get all officers who have tasks assigned
        cur.execute("""
            SELECT DISTINCT tm.assigned_officer
            FROM TASK_MASTER tm
            WHERE tm.assigned_officer IS NOT NULL
        """)
        officer_ids = [r[0] for r in cur.fetchall()]

    if not officer_ids:
        return {}

    scores = {}
    for oid in officer_ids:
        with db_cursor() as cur:
            # Farmer count for this officer
            cur.execute("""
                SELECT COUNT(DISTINCT tfm.farmer_id)
                FROM TASK_FARMER_MAPPING tfm
                JOIN TASK_MASTER tm ON tm.task_id = tfm.task_id
                WHERE tm.assigned_officer = ?
            """, oid)
            farmer_count = cur.fetchone()[0] or 0

            # Meeting KPIs
            cur.execute("""
                SELECT
                    COUNT(*) AS total_meetings,
                    SUM(CASE WHEN kcc = 1 THEN 1 ELSE 0 END) AS kcc_yes,
                    SUM(CASE WHEN cane_registration = 1 THEN 1 ELSE 0 END) AS cane_reg_yes,
                    SUM(CASE WHEN recovery = 1 THEN 1 ELSE 0 END) AS recovery_yes
                FROM TbL_TRN_Farmer_Meeting
                WHERE employee_id = ?
            """, oid)
            mk = cur.fetchone()

        total_meetings = mk[0] or 0
        if total_meetings > 0:
            kcc_pct = round((mk[1] or 0) / total_meetings * 100, 2)
            cane_reg_pct = round((mk[2] or 0) / total_meetings * 100, 2)
            recovery_pct = round((mk[3] or 0) / total_meetings * 100, 2)
        else:
            kcc_pct = cane_reg_pct = recovery_pct = 0

        meetings_pct = min(round(total_meetings / farmer_count * 100, 2), 100) if farmer_count > 0 else 0
        crushing_pct = cane_reg_pct  # proxy

        # Weighted: 40% Crushing + 20% Recovery + 20% Meetings + 20% KCC
        calc_score = round(crushing_pct * 0.4 + recovery_pct * 0.2 + meetings_pct * 0.2 + kcc_pct * 0.2, 2)
        scores[oid] = calc_score

    # Sort by score descending to assign ranks
    sorted_officers = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    total = len(sorted_officers)
    rankings = {}
    for rank_idx, (oid, score) in enumerate(sorted_officers, start=1):
        rankings[oid] = {"rank": rank_idx, "total": total, "score": score}

    return rankings


def generate_task_report(task_id: int, manager_id: int = 1):
    """
    Gathers data for a task, creates REPORT_MASTER & REPORT_DETAILS records,
    generates a detailed CSV snapshot with grades, rankings, trends, missed farmers.
    """
    with db_cursor() as cur:
        # 1. Fetch Task Header Details
        cur.execute("""
            SELECT 
                t.task_id, t.task_name, t.work_type, t.assigned_officer, 
                t.priority, t.start_date, t.end_date, t.status, t.remarks, t.created_at,
                ISNULL(u.first_name, '') + ' ' + ISNULL(u.last_name, '') AS officer_name,
                u.username AS officer_code,
                ISNULL(t.schedule_type, 'WEEKLY') AS schedule_type
            FROM TASK_MASTER t
            LEFT JOIN dbo.auth_user u ON u.id = t.assigned_officer
            WHERE t.task_id = ?
        """, task_id)
        task_row = cur.fetchone()
        if not task_row:
            print(f"Task ID {task_id} not found.")
            return None

        task_id = task_row[0]
        task_name = task_row[1]
        work_type = task_row[2]
        officer_id = task_row[3]
        priority = task_row[4]
        start_date = task_row[5]
        end_date = task_row[6]
        raw_status = task_row[7]
        remarks = task_row[8]
        created_at = task_row[9]
        officer_name = (task_row[10] or "").strip() or "Unknown"
        officer_code = task_row[11] or ""
        schedule_type = task_row[12] or "WEEKLY"

        dynamic_status = get_dynamic_status(raw_status, end_date)

        # 2. Fetch Mapped Farmers & details
        cur.execute("""
            SELECT 
                r.farmer_id,
                ISNULL(m.NameE, 'Unknown') AS farmer_name,
                ISNULL(v.Village_NameE, '') AS village,
                ISNULL(t_loc.Taluka_NameE, '') AS taluka,
                ISNULL(m.MobileNumber, '') AS mobile,
                r.status AS mapping_status
            FROM TASK_FARMER_MAPPING r
            LEFT JOIN dbo.TBL_MST_MASTER m ON m.code = r.farmer_id
            LEFT JOIN dbo.TBl_mst_village v ON v.Village_Code = m.Village_code
            LEFT JOIN dbo.TBl_mst_taluka t_loc ON t_loc.Taluka_Code = m.Talula_Code
            WHERE r.task_id = ?
        """, task_id)
        farmer_rows = cur.fetchall()

        farmers_list = []
        completed_farmers = 0
        pending_farmers = 0
        unique_villages = set()
        assigned_farmer_ids = set()

        for fr in farmer_rows:
            f_id, f_name, f_village, f_taluka, f_mobile, f_status = fr
            assigned_farmer_ids.add(f_id)
            farmers_list.append({
                "farmer_id": f_id,
                "name": f_name or "Unknown",
                "village": f_village,
                "taluka": f_taluka,
                "mobile": str(f_mobile) if f_mobile else "",
                "status": f_status
            })
            if f_status == 'COMPLETED':
                completed_farmers += 1
            else:
                pending_farmers += 1
            if f_village:
                unique_villages.add(f_village)

        farmer_count = len(farmers_list)
        village_count = len(unique_villages)
        completion_rate = round((completed_farmers / farmer_count * 100), 2) if farmer_count > 0 else 0.0

        # 3. Meeting KPIs — scoped to THIS task only (work_plan_id = task_id)
        #    Previously this was WHERE employee_id = ? which pulled the officer's
        #    entire lifetime history, inflating/mixing scores across tasks.
        cur.execute("""
            SELECT
                COUNT(*) AS total_meetings,
                SUM(CASE WHEN kcc = 1 THEN 1 ELSE 0 END) AS kcc_yes,
                SUM(CASE WHEN canara_hnt = 1 THEN 1 ELSE 0 END) AS canara_yes,
                SUM(CASE WHEN sangola_hnt = 1 THEN 1 ELSE 0 END) AS sangola_yes,
                SUM(CASE WHEN cane_registration = 1 THEN 1 ELSE 0 END) AS cane_reg_yes,
                SUM(CASE WHEN recovery = 1 THEN 1 ELSE 0 END) AS recovery_yes,
                SUM(CASE WHEN vehicle_agreement = 1 THEN 1 ELSE 0 END) AS vehicle_yes
            FROM TbL_TRN_Farmer_Meeting
            WHERE employee_id = ? AND work_plan_id = ?
        """, officer_id, task_id)
        mk = cur.fetchone()

        total_meetings = mk[0] or 0
        kcc_pct = round((mk[1] or 0) / total_meetings * 100, 2) if total_meetings else 0
        canara_pct = round((mk[2] or 0) / total_meetings * 100, 2) if total_meetings else 0
        sangola_pct = round((mk[3] or 0) / total_meetings * 100, 2) if total_meetings else 0
        cane_reg_pct = round((mk[4] or 0) / total_meetings * 100, 2) if total_meetings else 0
        recovery_pct = round((mk[5] or 0) / total_meetings * 100, 2) if total_meetings else 0
        vehicle_pct = round((mk[6] or 0) / total_meetings * 100, 2) if total_meetings else 0

        meetings_coverage_pct = min(round(total_meetings / farmer_count * 100, 2), 100) if farmer_count > 0 else 0
        crushing_pct = cane_reg_pct  # proxy

        # Weighted score: 40% Crushing + 20% Recovery + 20% Meetings + 20% KCC
        calculated_score = round(
            crushing_pct * 0.4 + recovery_pct * 0.2 + meetings_coverage_pct * 0.2 + kcc_pct * 0.2, 2
        )
        grade = compute_grade(calculated_score)

        # 4. Missed Farmers — farmers assigned to THIS task with no meeting recorded for it
        #    Previously this was WHERE employee_id = ? which marked a farmer as "met"
        #    even if the meeting was for a completely different task.
        cur.execute("""
            SELECT DISTINCT farmer_code
            FROM TbL_TRN_Farmer_Meeting
            WHERE employee_id = ? AND work_plan_id = ?
        """, officer_id, task_id)
        met_farmer_ids = set(r[0] for r in cur.fetchall())
        missed_farmers = []
        for f in farmers_list:
            if f["farmer_id"] not in met_farmer_ids:
                missed_farmers.append({
                    "farmer_id": f["farmer_id"],
                    "name": f["name"],
                    "village": f["village"]
                })

        # 5. Previous period score (from the last REPORT_DETAILS for this officer)
        cur.execute("""
            SELECT TOP 1 calculated_score
            FROM REPORT_DETAILS rd
            JOIN REPORT_MASTER rm ON rm.report_id = rd.report_id
            WHERE rd.officer_id = ? AND rd.calculated_score IS NOT NULL
            ORDER BY rm.generated_date DESC
        """, officer_id)
        prev_row = cur.fetchone()
        prev_score = float(prev_row[0]) if prev_row and prev_row[0] is not None else None
        score_trend = round(calculated_score - prev_score, 2) if prev_score is not None else None

    # 6. Officer Rankings
    rankings = compute_officer_rankings()
    rank_info = rankings.get(officer_id, {"rank": 0, "total": 0, "score": 0})

    # 7. Build snapshot JSON
    snapshot_data = {
        "task": {
            "task_id": task_id,
            "task_name": task_name,
            "work_type": work_type,
            "priority": priority,
            "start_date": str(start_date) if start_date else "",
            "end_date": str(end_date) if end_date else "",
            "status": dynamic_status,
            "schedule_type": schedule_type,
            "remarks": remarks or "",
            "officer_name": officer_name,
            "officer_code": officer_code
        },
        "farmers": farmers_list,
        "metrics": {
            "farmer_count": farmer_count,
            "village_count": village_count,
            "completed_farmers": completed_farmers,
            "pending_farmers": pending_farmers,
            "completion_rate": completion_rate
        },
        "kpis": {
            "kcc_pct": kcc_pct,
            "canara_pct": canara_pct,
            "sangola_pct": sangola_pct,
            "cane_reg_pct": cane_reg_pct,
            "recovery_pct": recovery_pct,
            "vehicle_pct": vehicle_pct,
            "meetings_coverage_pct": meetings_coverage_pct,
            "crushing_pct": crushing_pct,
            "calculated_score": calculated_score,
            "grade": grade,
        },
        "ranking": {
            "rank": rank_info["rank"],
            "total_officers": rank_info["total"],
        },
        "trend": {
            "prev_score": prev_score,
            "current_score": calculated_score,
            "change": score_trend,
        },
        "missed_farmers": missed_farmers,
    }
    snapshot_json = json.dumps(snapshot_data)

    # 8. Create Excel file
    now = datetime.datetime.now()
    year_str = now.strftime("%Y")
    month_str = now.strftime("%m")
    relative_path = os.path.join(year_str, month_str, f"task_{task_id}_report.xlsx")
    absolute_dir = os.path.join(REPORTS_DIR, year_str, month_str)
    os.makedirs(absolute_dir, exist_ok=True)
    absolute_filepath = os.path.join(absolute_dir, f"task_{task_id}_report.xlsx")

    meeting_rows = []
    if officer_id:
        with db_cursor() as cur:
            cur.execute("""
                SELECT
                    tfm.farmer_id,
                    ISNULL(m.NameE, 'Unknown') AS farmer_name,
                    ISNULL(v.Village_NameE, '') AS village,
                    fm.kcc, fm.kcc_reason,
                    fm.canara_hnt, fm.canara_reason,
                    fm.sangola_hnt, fm.sangola_reason,
                    fm.cane_registration, fm.cane_registration_remark,
                    fm.recovery, fm.recovery_reason,
                    fm.vehicle_agreement, fm.vehicle_reason,
                    fm.expected_tonnage,
                    fm.cane_development,
                    fm.feedback,
                    fm.remark,
                    CONVERT(VARCHAR, fm.created_at, 106) AS meeting_date,
                    fm.is_working_vehicle, fm.vehicle_working_reason
                FROM TASK_FARMER_MAPPING tfm
                LEFT JOIN dbo.TBL_MST_MASTER m ON m.code = tfm.farmer_id
                LEFT JOIN dbo.TBl_mst_village v ON v.Village_Code = m.Village_code
                LEFT JOIN TbL_TRN_Farmer_Meeting fm ON fm.farmer_code = tfm.farmer_id
                    AND fm.employee_id = ? AND fm.work_plan_id = ?
                WHERE tfm.task_id = ?
            """, officer_id, task_id, task_id)
            meeting_rows = cur.fetchall()

    with db_cursor() as cur:
        cur.execute("""
            SELECT 
                CONVERT(VARCHAR, al.timestamp, 120) AS timestamp_str,
                al.action,
                al.remarks,
                ISNULL(u.first_name,'') + ' ' + ISNULL(u.last_name,'') AS officer_name
            FROM TASK_ACTIVITY_LOG al
            LEFT JOIN dbo.auth_user u ON u.id = al.officer
            WHERE al.task_id = ?
            ORDER BY al.timestamp DESC
        """, task_id)
        history_rows = cur.fetchall()

    wb = build_excel_report(snapshot_data, meeting_rows, history_rows)
    wb.save(absolute_filepath)

    # 9. Insert into DB
    with db_cursor() as cur:
        report_type = f"{work_type or schedule_type} Performance Report"

        cur.execute("""
            INSERT INTO REPORT_MASTER (task_id, report_type, generated_by, status, file_path)
            OUTPUT INSERTED.report_id
            VALUES (?, ?, ?, 'COMPLETE', ?)
        """, task_id, report_type, manager_id, relative_path.replace("\\", "/"))

        report_id = cur.fetchone()[0]

        cur.execute("""
            INSERT INTO REPORT_DETAILS (
                report_id, officer_id, farmer_count, village_count, completion_rate,
                report_status, task_status, assigned_date, end_date, 
                completed_farmers, pending_farmers, generated_snapshot_json, remarks,
                performance_grade, calculated_score, officer_rank, total_officers,
                kcc_pct, canara_pct, sangola_pct, cane_reg_pct, recovery_pct,
                vehicle_pct, meetings_coverage_pct, prev_score, score_trend,
                missed_farmers_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        report_id, officer_id, farmer_count, village_count, completion_rate,
        'COMPLETE', dynamic_status, start_date, end_date,
        completed_farmers, pending_farmers, snapshot_json, remarks,
        grade, calculated_score, rank_info["rank"], rank_info["total"],
        kcc_pct, canara_pct, sangola_pct, cane_reg_pct, recovery_pct,
        vehicle_pct, meetings_coverage_pct, prev_score, score_trend,
        json.dumps(missed_farmers))

    print(f"Report ID {report_id} generated — Grade: {grade}, Score: {calculated_score}%, Rank: #{rank_info['rank']}/{rank_info['total']}")
    return report_id


def check_and_generate_eligible_reports(manager_id: int = 1):
    """
    Scans for tasks that are either completed OR past their end date (overdue),
    checks if they have a generated report in REPORT_MASTER, and if not, generates it.
    """
    with db_cursor() as cur:
        cur.execute("""
            SELECT t.task_id
            FROM TASK_MASTER t
            LEFT JOIN REPORT_MASTER r ON r.task_id = t.task_id
            WHERE r.report_id IS NULL
              AND (
                  t.status = 'COMPLETED'
                  OR (t.end_date < CAST(GETDATE() AS DATE) AND t.status NOT IN ('COMPLETED', 'CANCELLED'))
              )
        """)
        rows = cur.fetchall()

    task_ids = [r[0] for r in rows]
    generated = 0
    for tid in task_ids:
        try:
            result = generate_task_report(tid, manager_id)
            if result:
                generated += 1
        except Exception as e:
            print(f"Failed to generate report for Task ID {tid}: {e}")

    print(f"Auto-generation complete: {generated} new reports from {len(task_ids)} eligible tasks")
    return generated
